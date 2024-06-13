# Función para descargar las tablas en formato xlsx
def download_excel(df, button_text, name_download_data):
    import io # Descargar la tabla en xlsx
    excel_buffer = io.BytesIO()# Crear un objeto StringIO para almacenar los datos del archivo Excel en memoria
    df.to_excel(excel_buffer, index=False)# Guardar el DataFrame en un archivo Excel en el buffer
    # Obtener los bytes del buffer
    excel_buffer.seek(0)
    excel_binary = excel_buffer.getvalue()
    # Descargar el archivo Excel
    st.download_button(button_text, data=excel_binary, file_name=name_download_data, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# 1° Ingreso de módulos
import pandas as pd #pip install pandas
import plotly.express as px #pip install plotly-express
import streamlit as st #pip install streamlit
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.subplots as sp
import openpyxl

from datetime import datetime
import sys
import math
import folium
from streamlit_folium import st_folium
from folium.plugins import MiniMap

# 2° Ingreso de valores
    #2.1° Ingreso de valores genéricos - PRESENTACIÓN
title_page_web='Recloser' #Título del Dashboard
title_portada='🖥️ Recloser|Respuesta|Comunicación' #Título del Dashboard
name_empresa='Empresa Electrocentro S.A.' #Título de la empresa

lista_pestañas=["Presentación","1- Recloser instalados.", "2- Recloser comunicación.","3- Puntos críticos."] # Lista de las pestañas
name_excel='Registros.xlsx' # Base de datos
name_ciudades="Ciudades.xlsx" # Coordenadas de los departamentos.
imagen_path = "imagen.jpg"  # Ajusta la ruta de la imagen según sea necesario

    # 2.2° PESTAÑA (01)
lista_recloser=["NOJA","NOJA Power","SCHNEIDER","JINKWANG","ENTEC","S&C","ABB","SEL"] # Recloser instalados (para el conteo).
# URL de Google Maps
url_input = "https://www.google.com/maps/d/u/0/viewer?mid=1jDCOXn4Su3ub1LHtoZyHbpffU_0ZwdA&ll=-11.344651744765466%2C-73.25285471281072&z=7"

selected_columns = ['AMT',
                    'MARCA',
                    'Controlador',
                    'SECC. GIS NUEVO',
                    'UBICACIÓN',
                    'OPERADOR INSTALADO',
                    'IP DEL CHIP',
                    'Rpta actual',
                    'Comunicación actual'] # Columnas mostradas por defecto en la BD (pestaña 1).

    # 2.3° PESTAÑA (02)
selected_columns_comunicacion = ['Fecha',
                    'AMT',
                    'MARCA',
                    'Controlador',
                    'Codigo SCADA Actual',
                    'SECC. GIS NUEVO',
                    'OPERADOR INSTALADO',
                    'IP DEL CHIP',
                    'Rpta actual',
                    'Comunicación actual'] # Columnas mostradas por defecto en la BD (pestaña 2)

si_rpta_color='#3D2B8E' # Color de la barra, del día con respuesta.
si_com_color='#badb73'  # Color de la barra, del día con comunicación.
no_color='#ffaaa6'      # Color de la barra, del día sin respuesta o comunicación.

    # 2.4° PESTAÑA (03)
selected_columns_criticos = ['AMT',
                    'MARCA',
                    'Controlador',
                    'Codigo SCADA Actual',
                    'UBICACIÓN',
                    'OPERADOR INSTALADO',
                    'IP DEL CHIP',
                    'N° días incomunicados'] # Columnas mostradas por defecto en la BD (pestaña 3).
 
#********************************************************************************************************
st.set_page_config(page_title = title_page_web, #Nombre de la pagina, sale arriba cuando se carga streamlit
                    page_icon = '⚡', # https://www.webfx.com/tools/emoji-cheat-sheet/
                    layout="wide")

selected_tab = st.sidebar.radio("Visualización: ", lista_pestañas) # Pestañas en el menú lateral.
st.sidebar.markdown("---")# Insertar una línea horizontal

if selected_tab == lista_pestañas[0]:
    # 3° Nombres de la página web.
    col1, col2 = st.columns([7, 1])  # Usar proporciones para especificar el ancho relativo de cada columna

    col1.title(title_portada)
    col1.subheader(name_empresa)
    col1.subheader("_Elaborado por_: :blue[S.D.C.A] 👷")#, divider='rainbow')

    col2.image(imagen_path, use_column_width=True)
    st.markdown('---')
else:
    # # Poner logo.
    # col1, col2 = st.columns([7, 1])  # Usar proporciones para especificar el ancho relativo de cada columna
    # col2.image(imagen_path, use_column_width=True)
    # st.markdown('---')

    if selected_tab == lista_pestañas[1]:

        # Abre el libro de Excel "Registros.xlsx"
        workbook = openpyxl.load_workbook(name_excel)

        sheet_names = []
        for sheet in workbook.sheetnames:# Recorre todos los nombres del libro
            if not(sheet in ["SELECTORES","PLANTILLA","BDGeneral"]):
                sheet_names.append(sheet) # Lista que almacenará los nombres de las hojas
        workbook.close()# Cierra el libro

        # Selección de fecha
        try:
            min_date = datetime.strptime(sheet_names[0], "%d-%m-%Y")
            selected_date = st.sidebar.date_input(f"Selecciona una fecha a partir de {sheet_names[0]}:", min_value=min_date)
            hoja_excel = selected_date.strftime("%d-%m-%Y")

#!!!!!!!!!!!!!!!!11
            # 4° Lectura de los datos de la hoja excel seleccionada.
            df = pd.read_excel(name_excel,sheet_name = hoja_excel)
            
                #4.1° Filtrar los datos necesarios.
            df_filtro_Marca=df[df['MARCA'].isin(lista_recloser)]
            df_filtro_Marca["MARCA"]=df_filtro_Marca["MARCA"].replace("NOJA Power", "NOJA")
            df_filtro_Marca["MARCA"]=df_filtro_Marca["MARCA"].replace("Entec", "ENTEC")
                                                                                                                #El "\" indica un salto.
            condicion_filtrar= ~((df_filtro_Marca["MARCA"] == "S&C") & (df_filtro_Marca["SECC. GIS NUEVO"] == "--")) | \
                ~((df_filtro_Marca['MARCA'] == 'ABB') & (df_filtro_Marca['Controlador'] != "PCD2000R")) | \
                ~((df_filtro_Marca['MARCA'] == 'SEL') & (df_filtro_Marca['Controlador'] != "SEL-351R"))
            
            df_filtrado = df_filtro_Marca[condicion_filtrar]
            df_filtro_Marca = df_filtro_Marca.reset_index(drop=True) # Dataframe filtrado, que se usará para los siguientes filtros.

        except Exception as e:
            st.error("No hay registro para la fecha seleccionada, seleccione otra fecha.")
            sys.exit() # En caso de existir un error. Terminar de ejecutar el programa.
      
        try:
#!!!!!!!!!!!!!!!!22
        # 5° Creación de tablas
            # 5.1 Creación de filtros
            st.sidebar.markdown("---")
            select_all = st.sidebar.checkbox("""Considerar todos los registros:
                                                \n-Departamentos (DPTOS)
                                                \n-Unidades de Negocio (UN)
                                                \n-Subestaciones Eléctricas (SE)
                                            """) #Casilla de verificación para seleccionar todos los filtros.
            if select_all:
                st.sidebar.warning('Si desea filtrar, entonces deberá de desmarcar la casilla de verificación.')
                dpto = df_filtrado['DEPARTAMENTO'].unique()
                unidad_negocio = df_filtrado['UNIDAD DE NEGOCIO'].unique()
                se = df_filtrado['SUBESTACION'].unique()
                operador = df_filtrado['OPERADOR INSTALADO'].unique()
                amt = df_filtrado['AMT'].unique()

            else:
                # Filtro de Departamento
                st.sidebar.markdown("---")
                st.sidebar.header("Departamento:")
                select_all_dpto = st.sidebar.checkbox("Marcar todos los DPTOS.") # Casilla de verificación para seleccionar todos los filtros.
                lista_dpto=sorted(df_filtrado['DEPARTAMENTO'].unique().tolist()) # Lista que tiene a las opciones ordenadas

                if select_all_dpto:
                    st.sidebar.warning("Seleccionaste todos los DPTOS.")
                    dpto = lista_dpto
                else:
                    dpto = st.sidebar.multiselect(
                        "Filtrar dptos",
                        options=lista_dpto,
                        default=[],
                    )
                
                # Filtro de Unidad de Negocio (se habilita en función de la selección de Departamento)
                st.sidebar.header("Unidad de Negocio:")
                select_all_UN= st.sidebar.checkbox("Marcar todas las UN.")#Casilla de verificación para seleccionar todos los filtros.
                un_options = df_filtrado[df_filtrado['DEPARTAMENTO'].isin(dpto)]['UNIDAD DE NEGOCIO'].unique() # Filtro las UN por los dptos seleccionados.
                lista_UN_options=sorted(un_options.tolist())

                if select_all_UN:
                    st.sidebar.warning("Seleccionaste todas las UN.")
                    unidad_negocio = lista_UN_options
                else:
                    unidad_negocio = st.sidebar.multiselect(
                        "Filtrar UN:",
                        options=lista_UN_options, # Opciones, incluido a las UN ordenadas.
                        default=[],
                    )

                # Filtro de Subestación (se habilita en función de la selección de Unidad de Negocio)
                st.sidebar.header("Subestación Eléctrica")
                select_all_SE= st.sidebar.checkbox("Marcar todas las SE.")#Casilla de verificación para seleccionar todos los filtros.
                se_options = df_filtrado[
                    (df_filtrado['DEPARTAMENTO'].isin(dpto)) &
                    (df_filtrado['UNIDAD DE NEGOCIO'].isin(unidad_negocio))
                    ]['SUBESTACION'].unique()
                lista_se_options=sorted(se_options.tolist())

                if select_all_SE:
                    st.sidebar.warning("Seleccionaste todas las SE.")
                    se = lista_se_options
                else:
                    se = st.sidebar.multiselect(
                        "Filtrar SE:",
                        options=lista_se_options,
                        default=[],
                    )

                # Filtro de Operador (se habilita en función de la selección de Subestación)
                operador_options = df_filtrado[
                    (df_filtrado['DEPARTAMENTO'].isin(dpto)) &
                    (df_filtrado['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                    (df_filtrado['SUBESTACION'].isin(se))
                ]['OPERADOR INSTALADO'].unique()
                operador=sorted(operador_options.tolist())
                    
                # Filtro de Alimentador (AMT) - Filtrado en base a todas las selecciones anteriores
                amt_options = df_filtrado[
                    (df_filtrado['DEPARTAMENTO'].isin(dpto)) &
                    (df_filtrado['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                    (df_filtrado['SUBESTACION'].isin(se)) &
                    (df_filtrado['OPERADOR INSTALADO'].isin(operador))
                ]['AMT'].unique()
                amt=sorted(amt_options.tolist())
                
            # 5.2 Filtra el DataFrame en función de las selecciones:
            filtered_df = df_filtrado[
                (df_filtrado['DEPARTAMENTO'].isin(dpto)) &
                (df_filtrado['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                (df_filtrado['SUBESTACION'].isin(se)) &
                (df_filtrado['OPERADOR INSTALADO'].isin(operador)) &
                (df_filtrado['AMT'].isin(amt))
            ]

            # 5.3 Muestra la tabla con los datos filtrados
            # st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>TABLA DE DATOS </p>", unsafe_allow_html=True)
            # Mostrar la base de datos
            st.sidebar.markdown("---")
            st.sidebar.header("BASE DE DATOS")
            selected_columns_difference=[col for col in filtered_df.columns if col not in selected_columns] # Muestra de los demás campos.
            
            select_all_field = st.sidebar.checkbox("Mostrar todos los CAMPOS.") # Casilla de verificación para seleccionar todos los filtros.
            lista_field=sorted(selected_columns_difference) # Lista que tiene a las opciones ordenadas

            if select_all_field:
                st.sidebar.warning("Se mostrará todos los CAMPOS de la Base de Datos.")
                fields = lista_field

            else:
                fields = st.sidebar.multiselect(
                    "Filtrar campos",
                    options=lista_field,
                    default=[],
                )
            
            # Sacar el orden de la Base de Datos.
            Laux_orden=[]
            lista_orden_field=["AMT", "MARCA","SECC. GIS NUEVO"] # Registros ordenados por defecto.
            for elemento_field in fields:
                if elemento_field in ['DEPARTAMENTO', 'UNIDAD DE NEGOCIO', 'SUBESTACION']:
                    Laux_orden.append(elemento_field)
            Laux_orden.extend(lista_orden_field)
            lista_orden_field=Laux_orden
            
            st.sidebar.markdown("---") #separador
            
            selected_columns=selected_columns+fields # Lista con los campos a mostrar.
            df_base_data = filtered_df[filtered_df.columns.intersection(selected_columns)]# Ordenar las columnas en base a la BD original.
            df_base_data = df_base_data.sort_values(by=lista_orden_field) # Ordenar la base de datos en base al orden definido
            df_base_data = df_base_data.reset_index(drop=True)# Reiniciar la enumeración del DataFrame
            df_base_data.index = df_base_data.index + 1 #Hacer que la llave primaria inicie en "1" (INDEX).
            
            # st.write(df_base_data) # Imprimir tabla         
            # download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)

            # 6° Impresión de los KPI's:

            # Impresión de KPIs => Conteo del total de recloser
            list_aux_marcas,list_aux_conteo=[],[]
            list_aux_si_rpta,list_aux_si_com,list_aux_no_rpta,list_aux_no_com=[],[],[],[]

                # OPCIÓN N°01:
                    # Filtrar Dataframe
            df_filtro_Marca=filtered_df[filtered_df['MARCA'].isin(lista_recloser)]
            df_filtro_Marca["MARCA"]=df_filtro_Marca["MARCA"].replace("NOJA Power", "NOJA")
            df_filtro_Marca["MARCA"]=df_filtro_Marca["MARCA"].replace("Entec", "ENTEC")

                                                                                                                #El "\" indica un salto.
            condicion_filtrar=((df_filtro_Marca["MARCA"] == "S&C") & (df_filtro_Marca["SECC. GIS NUEVO"] == "--")) | \
                ((df_filtro_Marca['MARCA'] == 'ABB') & (df_filtro_Marca['Controlador'] != "PCD2000R")) | \
                ((df_filtro_Marca['MARCA'] == 'SEL') & (df_filtro_Marca['Controlador'] != "SEL-351R"))
            
            df_filtro_Marca = df_filtro_Marca[~condicion_filtrar] # Eliminar los registros que cumplan con la condición.
            df_filtro_Marca = df_filtro_Marca.reset_index(drop=True) # Dataframe filtrado, que se usará para los siguientes filtros.
            download_excel(df_filtro_Marca,"📥 Download Data","Recloser.xlsx") # Descargar en formato (xlsx)
                    # Conteo
            conteos_marcas = pd.DataFrame({'MARCA': df_filtro_Marca['MARCA'].value_counts().index, 'Total': df_filtro_Marca['MARCA'].value_counts().values})
            si_rpta = df_filtro_Marca.groupby('MARCA')['Rpta actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='si_rpta')
            no_rpta = df_filtro_Marca.groupby('MARCA')['Rpta actual'].apply(lambda x: (x == 'No').sum()).reset_index(name='no_rpta')
            si_comu = df_filtro_Marca.groupby('MARCA')['Comunicación actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='si_comu')
            no_comu = df_filtro_Marca.groupby('MARCA').apply(lambda x: ((x['Rpta actual'] == 'Si') & (x['Comunicación actual'] == 'No')).sum()).reset_index(name='no_comu') # Contar los que no tienen comunicación, pero sí respuesta.

                    # Unimos los conteos, en base a las coincidencias que tengan los elementos de su campo "MARCA".
            # conteos_marcas = pd.merge(conteos_marcas, si_rpta, on='MARCA', how='left')
            # conteos_marcas = pd.merge(conteos_marcas, no_rpta, on='MARCA', how='left')
            # conteos_marcas = pd.merge(conteos_marcas, si_comu, on='MARCA', how='left')
            # conteos_marcas = pd.merge(conteos_marcas, no_comu, on='MARCA', how='left')
                        #Los 4 códigos de encima se pueden realizar en uno solo:
            conteos_marcas = pd.merge(pd.merge(pd.merge(pd.merge(conteos_marcas, si_rpta, on='MARCA', how='left'), no_rpta, on='MARCA', how='left'), si_comu, on='MARCA', how='left'), no_comu, on='MARCA', how='left')

                    # Sacamos la cantidad total, los que tienen o no respuesta/comunicación
            total_recloser, si_rpta_total, si_com_total, no_rpta_total, no_com_total = conteos_marcas[['Total', 'si_rpta', 'si_comu', 'no_rpta', 'no_comu']].sum()
            
            # # # #     # OPCIÓN N°02:
            # # # #         # Filtrar Dataframe y conteo
            # # # # conteo_NOJA = filtered_df.loc[filtered_df['MARCA'].isin(['NOJA', 'NOJA Power']), 'Controlador'].count() # Conteo para los NOJA.
            # # # # si_rpta_NOJA = filtered_df.loc[filtered_df['MARCA'].isin(['NOJA', 'NOJA Power']), 'Rpta actual'].eq("Si").sum() # Conteo para los NOJA que tienen respuesta.
            # # # # si_com_NOJA = filtered_df.loc[filtered_df['MARCA'].isin(['NOJA', 'NOJA Power']), 'Comunicación actual'].eq("Si").sum() # Conteo para los NOJA que tienen comunicación.
            # # # # no_rpta_NOJA = conteo_NOJA-si_rpta_NOJA # Conteo para los NOJA que no tienen respuesta.
            # # # # no_com_NOJA = si_rpta_NOJA-si_com_NOJA # Conteo para los NOJA que no tienen respuesta.
            
            # # # # conteo_Schneider = filtered_df.loc[filtered_df['MARCA']=='Schneider','Controlador'].count() # Conteo poara los SCHNEIDER
            # # # # si_rpta_Schneider = filtered_df.loc[filtered_df['MARCA']=='Schneider', 'Rpta actual'].eq("Si").sum() 
            # # # # si_com_Schneider = filtered_df.loc[filtered_df['MARCA']=='Schneider', 'Comunicación actual'].eq("Si").sum()
            # # # # no_rpta_Schneider = conteo_Schneider-si_rpta_Schneider
            # # # # no_com_Schneider = si_rpta_Schneider-si_com_Schneider

            # # # # conteo_JinkWang  = filtered_df.loc[filtered_df['MARCA']=='JinkWang','Controlador'].count() # Conteo poara los JinkWang
            # # # # si_rpta_JinkWang = filtered_df.loc[filtered_df['MARCA']=='JinkWang', 'Rpta actual'].eq("Si").sum() 
            # # # # si_com_JinkWang = filtered_df.loc[filtered_df['MARCA']=='JinkWang', 'Comunicación actual'].eq("Si").sum()
            # # # # no_rpta_JinkWang = conteo_JinkWang-si_rpta_JinkWang
            # # # # no_com_JinkWang = si_rpta_JinkWang-si_com_JinkWang

            # # # # conteo_ENTEC = filtered_df.loc[filtered_df['MARCA']=='ENTEC','Controlador'].count() # Conteo poara los ENTEC
            # # # # si_rpta_ENTEC = filtered_df.loc[filtered_df['MARCA']=='ENTEC', 'Rpta actual'].eq("Si").sum() 
            # # # # si_com_ENTEC = filtered_df.loc[filtered_df['MARCA']=='ENTEC', 'Comunicación actual'].eq("Si").sum()
            # # # # no_rpta_ENTEC = conteo_ENTEC-si_rpta_ENTEC
            # # # # no_com_ENTEC = si_rpta_ENTEC-si_com_ENTEC
            
            # # # # conteo_SC  = filtered_df.loc[filtered_df['MARCA'] == 'S&C', 'SECC. GIS NUEVO'].ne("--").sum() # Conteo para los S&C cuyo nombre de su Controlador es diferente a "--"
            # # # # si_rpta_SC = filtered_df.loc[(filtered_df['MARCA'] == 'S&C') & (filtered_df['SECC. GIS NUEVO'] != "--") & (filtered_df['Rpta actual'] == 'Si')].shape[0]
            # # # # si_com_SC = filtered_df.loc[(filtered_df['MARCA'] == 'S&C') & (filtered_df['SECC. GIS NUEVO'] != "--") & (filtered_df['Comunicación actual'] == 'Si')].shape[0]
            # # # # no_rpta_SC = conteo_SC-si_rpta_SC
            # # # # no_com_SC = si_rpta_SC-si_com_SC
            
            # # # # conteo_ABB = filtered_df.loc[filtered_df['MARCA']=='ABB','Controlador'].eq("PCD2000R").sum() # Conteo para los ABB cuyo nombre de su Controlador es igual a "PCD2000R"
            # # # # si_rpta_ABB = filtered_df.loc[(filtered_df['MARCA'] == 'ABB') & (filtered_df['Controlador'] != "PCD2000R") & (filtered_df['Rpta actual'] == 'Si')].shape[0]
            # # # # si_com_ABB = filtered_df.loc[(filtered_df['MARCA'] == 'ABB') & (filtered_df['Controlador'] != "PCD2000R") & (filtered_df['Comunicación actual'] == 'Si')].shape[0]
            # # # # no_rpta_ABB = conteo_ABB-si_rpta_ABB
            # # # # no_com_ABB = si_rpta_ABB-si_com_ABB

            # # # # conteo_SEL = filtered_df.loc[filtered_df['MARCA']=='SEL','Controlador'].eq("SEL-351R").sum() # Conteo para los SEL cuyo nombre de su Controlador es igual a "SEL-351R"
            # # # # si_rpta_SEL = filtered_df.loc[(filtered_df['MARCA'] == 'SEL') & (filtered_df['Controlador'] != "SEL-351R") & (filtered_df['Rpta actual'] == 'Si')].shape[0]
            # # # # si_com_SEL = filtered_df.loc[(filtered_df['MARCA'] == 'SEL') & (filtered_df['Controlador'] != "SEL-351R") & (filtered_df['Comunicación actual'] == 'Si')].shape[0]
            # # # # no_rpta_SEL = conteo_SEL-si_rpta_SEL
            # # # # no_com_SEL = si_rpta_SEL-si_com_SEL

            # # # #         # Creación del dataframe para el conteo de las Marcas.
            # # # # list_aux_marcas=['NOJA','Schneider','JinkWang','ENTEC','S&C','ABB','SEL']
            # # # # list_aux_conteo=[conteo_NOJA,conteo_Schneider,conteo_JinkWang,conteo_ENTEC,conteo_SC,conteo_ABB,conteo_SEL]
            # # # # list_aux_si_rpta=[si_rpta_NOJA,si_rpta_Schneider,si_rpta_JinkWang,si_rpta_ENTEC,si_rpta_SC,si_rpta_ABB,si_rpta_SEL]
            # # # # list_aux_si_com=[si_com_NOJA,si_com_Schneider,si_com_JinkWang,si_com_ENTEC,si_com_SC,si_com_ABB,si_com_SEL]
            # # # # list_aux_no_rpta=[no_rpta_NOJA,no_rpta_Schneider,no_rpta_JinkWang,no_rpta_ENTEC,no_rpta_SC,no_rpta_ABB,no_rpta_SEL]
            # # # # list_aux_no_com=[no_com_NOJA,no_com_Schneider,no_com_JinkWang,no_com_ENTEC,no_com_SC,no_com_ABB,no_com_SEL]

            # # # # #Creación del diccionario, para la tabla.
            # # # # diccionario_marca={
            # # # #     'Marca':list_aux_marcas,
            # # # #     'Total':list_aux_conteo,
            # # # #     'Si Rpta':list_aux_si_rpta,
            # # # #     "No Rpta":list_aux_no_rpta,
            # # # #     "Si Com":list_aux_si_com,
            # # # #     "No Com":list_aux_no_com
            # # # # }

            # # # # conteos_marcas=pd.DataFrame(diccionario_marca)
            # # # # conteos_marcas = conteos_marcas.sort_values(by='Total', ascending=False) # Ordenar en base al número de recloser.
            # # # # total_recloser=conteos_marcas['Total'].sum()
            
            # # # # si_rpta_total=sum(list_aux_si_rpta)
            # # # # si_com_total=sum(list_aux_si_com)
            # # # # no_rpta_total=sum(list_aux_no_rpta)
            # # # # no_com_total=sum(list_aux_no_com)


    # RECLOSER INSTALADOS ----------------------------------------------------------------------------------------------------------------------------------------------------------


            # col1, col2 = st.columns([3.5,2.5]) #Centrar el botón
            # with col1:
            #     st.markdown(f"<p style='font-size: 10px; text-align: right; font-weight: bold;'>Recloser instalados: {total_recloser}</p>", unsafe_allow_html=True)
            # with col2:
            #     st.markdown(f"[[ Ubicación de recloser en Google Maps ]]({url_input})")
        # 6° Guardar el gráfico de barras en la siguiente variable
            #6.1° Creación de gráficas por MARCA.
            
            # 6.1.1° Recloser instalados por marcas-----------------> RECLOSERS INSTALADOS
            # st.subheader("Total de recloser por marca")
            figMARCA1 = px.bar(conteos_marcas, x="MARCA", y="Total",
                        text=['{:,.0f} und.'.format(x) for x in conteos_marcas["Total"]],
                        template="seaborn")
                # Configuración para mostrar el texto encima de las barras y con tamaño 24
            figMARCA1.update_traces(textposition='outside', textfont_size=15, marker_line_color = 'black', marker_line_width = 2) # Config. de las etiquetas de las barras.
            figMARCA1.update_layout(xaxis=dict(tickangle=-45, tickfont=dict(size=15)),yaxis_range=[0, conteos_marcas['Total'].max()+30]) # Config. texto del eje "X"
            
            # st.plotly_chart(figMARCA1, use_container_width=True, height=200) # Mostrar gráfica
            
            # 6.1.2° Agrupar por 'UNIDAD DE NEGOCIO' => "Nro de recloser instalados"----------------->DIAGRAMA DE PASTEL
            # st.subheader("Porcentaje de recloser por marca")
                # GRAFICAR
            figMARCA2=px.pie(conteos_marcas,values='Total',hole=0.25)
            figMARCA2.update_traces(text=conteos_marcas['MARCA'], textposition='outside',textfont_size=15)
            # st.plotly_chart(figMARCA2,use_container_width=True) # Mostrar gráfica

            # 6.2° Creación de tablas
                # 6.2.1: Creación de la tabla de los diagramas
            conteos_marcas.reset_index(drop=True, inplace=True)
            conteos_marcas['Porcentaje (%)']=round(conteos_marcas['Total']/total_recloser*100,2)
            conteos_marcas.index = conteos_marcas.index + 1  # Hacer que la primera fila no sea "0"

            # # with st.expander("Marca_ViewData"):
            # #     st.write(conteos_marcas)
            # #     # st.write(conteos_marcas.style.background_gradient(cmap="Greens"))# Imprimir la tabla con escala de colores
            # #     download_excel(conteos_marcas,"📥 Download Data","Marca-DATA.xlsx") # Descargar en formato (xlsx)

            #**********************************************************************************************************************************************************
# GRÁFICA RESPUESTA/COMUNICACIÓN ----------------------------------------------------------------------------------------------------------------------------------------------------------
            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Respuesta de recloser: {si_rpta_total}/{total_recloser}</p>", unsafe_allow_html=True)
            # st.markdown(f"<p style='font-size: 22px;  text-align: Center'>Comunicación de recloser: {si_com_total}/{si_rpta_total}</p>", unsafe_allow_html=True)
            
            # Gráfico de velocímetro, con nro de respuestas y comunicaciones.
                # Graficar
            layout = go.Layout(
                grid = {'rows': 1, 'columns': 2, 'pattern': 'independent'}, # El pattern también es: independent ver argumentos en => https://plotly.com/python-api-reference/generated/plotly.graph_objects.Layout.html
                width = 600,  # Width in pixels
                height = 300  # Height in pixels
            )# Create the layout of the plot

            # First value
            figRC = go.Figure(layout=layout)
            # First value
            figRC.add_trace(go.Indicator(
                mode="number+delta+gauge",
                value=si_rpta_total,
                delta={'reference': math.ceil(total_recloser/2)},
                gauge={
                    'axis': {'visible': True, 'range': [None, total_recloser]},
                    'steps': [{'range': [0, si_rpta_total], 'color': "lightgray"}],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': math.ceil(total_recloser/2)
                    } # Línea de referencia
                },
                title={"text": f"Respuesta/{total_recloser}"},
                domain={'x': [0, 0.5], 'y': [0, 1]}
            ))
            # Second value
            figRC.add_trace(go.Indicator(
                mode="number+delta+gauge",
                value=si_com_total,
                delta={'reference': math.ceil(si_rpta_total/2)},
                gauge={
                    'axis': {'visible': True, 'range': [None, si_rpta_total]},
                    'steps': [{'range': [0, si_com_total], 'color': "lightgray"}],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': math.ceil(si_rpta_total/2)
                    } # Línea de referencia
                },
                title={"text": f"Comunicación/{si_rpta_total}"},
                domain={'x': [0.5, 1], 'y': [0, 1]}
            ))
            # st.plotly_chart(figRC)

            #******************************************************************************************************************************************************************************************************************
            #****************************************************************************************************************************************************************************************************************** 
            #6.3° RECLOSER POR DEPARTAMENTO ==> Usaremos el dataframe filtrado (df_filtro_Marca)
                # Conteo
            conteos_marcas_DPTO = pd.DataFrame({'DEPARTAMENTO': df_filtro_Marca['DEPARTAMENTO'].value_counts().index, 'Recloser instalados': df_filtro_Marca['DEPARTAMENTO'].value_counts().values})
            si_rpta = df_filtro_Marca.groupby('DEPARTAMENTO')['Rpta actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con respuesta')
            no_rpta = df_filtro_Marca.groupby('DEPARTAMENTO')['Rpta actual'].apply(lambda x: (x == 'No').sum()).reset_index(name='Recloser sin respuesta')
            si_comu = df_filtro_Marca.groupby('DEPARTAMENTO')['Comunicación actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con comunicación')
            no_comu = df_filtro_Marca.groupby('DEPARTAMENTO').apply(lambda x: ((x['Rpta actual'] == 'Si') & (x['Comunicación actual'] == 'No')).sum()).reset_index(name='Recloser sin comunicación') # Contar los que no tienen comunicación, pero sí respuesta.

                    # Unimos los conteos, en base a las coincidencias que tengan los elementos de su campo "DEPARTAMENTO".
            conteos_marcas_DPTO = pd.merge(pd.merge(pd.merge(pd.merge(conteos_marcas_DPTO, si_rpta, on='DEPARTAMENTO', how='left'), no_rpta, on='DEPARTAMENTO', how='left'), si_comu, on='DEPARTAMENTO', how='left'), no_comu, on='DEPARTAMENTO', how='left')
            
            conteos_marcas_DPTO = conteos_marcas_DPTO.sort_values(by='DEPARTAMENTO').reset_index(drop=True) #Ordenar por DPTO.
            conteos_marcas_DPTO.index = conteos_marcas_DPTO.index + 1 # Hacer que el índice comience en 1.

            #6.3° GRÁFICA POR DEPARTAMENTO ==> Usaremos el dataframe filtrado (df_filtro_Marca)
            df_folium=pd.read_excel(name_ciudades,sheet_name="Ciudades")
            CM_latitud,CM_longitud=-9.189967,-75.015152 # Punto de inicio

            map_folium = folium.Map(location=[CM_latitud,CM_longitud], zoom_start=6) # Creación del mapa folium centrado en un punto.
            
            for index,row in conteos_marcas_DPTO.iterrows():
                Dep=row["DEPARTAMENTO"]
                indice_dpto = df_folium.index[df_folium['DEPARTAMENTO'] == Dep].tolist()# Obtener el índice donde el valor en la columna "Dpto" es "Arequipa"

                # Sacar las coordenadas y el color de los dptos seleccionados del excel:
                lat=df_folium.at[indice_dpto[0], 'LATITUD']
                long=df_folium.at[indice_dpto[0], 'LONGITUD']
                name_color=df_folium.at[indice_dpto[0], 'COLOR']

                Rec_total_DPTO=row['Recloser instalados'] 
                etiqueta_cuadro=f'{Dep}\n(cantidad:\n {Rec_total_DPTO} Rec.)'

                folium.Marker(location=[lat,long], icon=folium.Icon(color=name_color), popup=etiqueta_cuadro, tooltip=Dep).add_to(map_folium) # El popup es el cuadro que sale en el mapa.
                folium.Circle(location=[lat,long], color=name_color,radius=4,weight=20,fill_opacity=0.5).add_to(map_folium)

                    #Agregar minimapa
            folium.TileLayer().add_to(map_folium)
            map_mini=MiniMap()
            map_folium.add_child(map_mini)

            # st_data = st_folium(map_folium, width=725)# Mostrar el mapa de Folium en Streamlit

            # # Mostrar tabla
            # with st.expander("DPTO_ViewData"):
            #     st.write(conteos_marcas_DPTO)
            #     download_excel(conteos_marcas_DPTO,"📥 Download Data","Departamento-DATA.xlsx") # Descargar en formato (xlsx)

            #******************************************************************************************************************************************************************************************************************
            #****************************************************************************************************************************************************************************************************************** 
            #6.4° RECLOSER POR UNIDAD DE NEGOCIO ==> Usaremos el dataframe filtrado (df_filtro_Marca)
                # Conteo
            conteos_marcas_UN = pd.DataFrame({'UNIDAD DE NEGOCIO': df_filtro_Marca['UNIDAD DE NEGOCIO'].value_counts().index, 'Recloser instalados': df_filtro_Marca['UNIDAD DE NEGOCIO'].value_counts().values})
            si_rpta = df_filtro_Marca.groupby('UNIDAD DE NEGOCIO')['Rpta actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con respuesta')
            no_rpta = df_filtro_Marca.groupby('UNIDAD DE NEGOCIO')['Rpta actual'].apply(lambda x: (x == 'No').sum()).reset_index(name='Recloser sin respuesta')
            si_comu = df_filtro_Marca.groupby('UNIDAD DE NEGOCIO')['Comunicación actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con comunicación')
            no_comu = df_filtro_Marca.groupby('UNIDAD DE NEGOCIO').apply(lambda x: ((x['Rpta actual'] == 'Si') & (x['Comunicación actual'] == 'No')).sum()).reset_index(name='Recloser sin comunicación') # Contar los que no tienen comunicación, pero sí respuesta.

                    # Unimos los conteos, en base a las coincidencias que tengan los elementos de su campo "UNIDAD DE NEGOCIO".
            conteos_marcas_UN = pd.merge(pd.merge(pd.merge(pd.merge(conteos_marcas_UN, si_rpta, on='UNIDAD DE NEGOCIO', how='left'), no_rpta, on='UNIDAD DE NEGOCIO', how='left'), si_comu, on='UNIDAD DE NEGOCIO', how='left'), no_comu, on='UNIDAD DE NEGOCIO', how='left')
            
            conteos_marcas_UN = conteos_marcas_UN.sort_values(by='UNIDAD DE NEGOCIO').reset_index(drop=True) #Ordenar por UN
            conteos_marcas_UN.index = conteos_marcas_UN.index + 1 # Hacer que el índice comience en 1.

# GRÁFICA UN ----------------------------------------------------------------------------------------------------------------------------------------------------------
            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Recloser por Unidad de Negocio</p>", unsafe_allow_html=True)            
            # # OPCION 1: RADAR-CHART
            fig_UN = go.Figure()
            # Definir los datos de las curvas y sus nombres: Para cerrar el diagrama se deberá de repetir los primeros valores de los radios y del ángulo.
            curvas=[]
            ejes=['Con Res','Con Com','Sin Res','Sin Comu','Con Res'] # Al final duplicamos el primer eje, para que el diagrama sea cerrado.
            for index,row in conteos_marcas_UN.iterrows():
                Sres,Nres,Scom,Ncom=row['Recloser con respuesta'],row['Recloser con comunicación'],row['Recloser sin respuesta'],row['Recloser sin comunicación']
                name_UN=row['UNIDAD DE NEGOCIO']

                lista_radar_chart=[Sres,Nres,Scom,Ncom,Sres] # Al final duplicamos el primer valor, para que el diagrama sea cerrado.
                curvas.append({"r":lista_radar_chart, "theta":ejes,"nombre":name_UN})

            # Número de filas y columnas de la gráfica del radar.
            n_fig_UUNN=len(conteos_marcas_UN) # Graficará 1 UUNN por figura
            nro_filas_radar=1 # Número de filas "SE"

            if n_fig_UUNN>4:
                nro_filas_radar=2

            nro_columnas_radar=n_fig_UUNN//nro_filas_radar # Cociente = n° de columnas            

            # Añadir las figuras un arreglos.
            
            fig_UN = sp.make_subplots(rows=nro_filas_radar, cols=nro_columnas_radar, specs=[[{'type': 'polar'}]*nro_columnas_radar]*nro_filas_radar)

            for idx,curva in enumerate(curvas):
                posicion_fila=idx//nro_columnas_radar+1
                posicion_columna=idx%nro_columnas_radar+1
                
                fig_UN.add_trace(go.Scatterpolar(
                    r=curva["r"],
                    theta=curva["theta"],
                    fill=None,  # No mostrar relleno, también suele tener el parámetro de "toself" para que el relleno sea de igual color que el contorno.
                    name=curva["nombre"]
                ), row=posicion_fila, col=posicion_columna)  # Esto puede variar según la posición en la subfigura


            # Actualizar diseño del gráfico
            fig_UN.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True
                    ),
                ),
                showlegend=True,  # Mostrar leyenda
                template="plotly_dark"  # Fondo oscuro de la figura
            )
            # st.plotly_chart(fig_UN)

            # # # OPCION 2: GRAFICO DE BARRAS HORIZONTAL (Graficar con px.bar es diferente a graficar con go.bar)
            # # fig_UN = px.bar(conteos_marcas_UN, 
            # #                     x=['Recloser instalados','Recloser con respuesta','Recloser con comunicación'], 
            # #                     y=conteos_marcas_UN['UNIDAD DE NEGOCIO'],
            # #                     orientation="h", 
            # #                     color_discrete_sequence=["#C2BC18", "#FAA632", '#13B3C1'],
            # #                     opacity=[1], #Opacidad
            # #                     template='plotly_white')
            
            # # fig_UN.update_layout(
            # #     title="",
            # #     plot_bgcolor="rgba(0,0,0,0)",
            # #     yaxis_title='Unidad de Negocio',
            # #     xaxis_title='Cantidad de recloser',
            # #     title_x=0,
            # #     autosize=True,
            # #     height=450,
            # #     barmode='overlay'# Superponer las barras en lugar de apilarlas
            # # )

            # # # Configurar el tamaño del gráfico
            # # fig_UN.update_layout(height=450)
            # # fig_UN.update_xaxes(range=[0, conteos_marcas_UN['Recloser instalados'].max()+5],
            # #                         tickvals=list(range(0, conteos_marcas_UN['Recloser instalados'].max()+5, 20)))

            # # fig_UN.update_traces(textposition='outside', textfont_size=15) 
            
            # # # # Mostrar el gráfico en Streamlit
            # # # st.plotly_chart(fig_UN, use_container_width=True)
            
            # # Mostrar tabla
            # with st.expander("UN_ViewData"):
            #     st.write(conteos_marcas_UN)
            #     download_excel(conteos_marcas_UN,"📥 Download Data","Unidad de Negocio-DATA.xlsx") # Descargar en formato (xlsx)
            
            # st.markdown("---") #separador
    #******************************************************************************************************************************************************************************************************************
    #******************************************************************************************************************************************************************************************************************
            # 6.5° Agrupar por 'SUBESTACIÓN' => "Nro de respuestas y comunicación de los recloser"
                # Conteo
            conteos_marcas_SE = pd.DataFrame({'SUBESTACION': df_filtro_Marca['SUBESTACION'].value_counts().index, 'Recloser instalados': df_filtro_Marca['SUBESTACION'].value_counts().values})
            si_rpta = df_filtro_Marca.groupby('SUBESTACION')['Rpta actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con respuesta')
            no_rpta = df_filtro_Marca.groupby('SUBESTACION')['Rpta actual'].apply(lambda x: (x == 'No').sum()).reset_index(name='Recloser sin respuesta')
            si_comu = df_filtro_Marca.groupby('SUBESTACION')['Comunicación actual'].apply(lambda x: (x == 'Si').sum()).reset_index(name='Recloser con comunicación')
            no_comu = df_filtro_Marca.groupby('SUBESTACION').apply(lambda x: ((x['Rpta actual'] == 'Si') & (x['Comunicación actual'] == 'No')).sum()).reset_index(name='Recloser sin comunicación') # Contar los que no tienen comunicación, pero sí respuesta.

                    # Unimos los conteos, en base a las coincidencias que tengan los elementos de su campo "SUBESTACION".
            conteos_marcas_SE = pd.merge(pd.merge(pd.merge(pd.merge(conteos_marcas_SE, si_rpta, on='SUBESTACION', how='left'), no_rpta, on='SUBESTACION', how='left'), si_comu, on='SUBESTACION', how='left'), no_comu, on='SUBESTACION', how='left')
            
            conteos_marcas_SE = conteos_marcas_SE.sort_values(by='SUBESTACION').reset_index(drop=True) #Ordenar por nombre
            conteos_marcas_SE.index = conteos_marcas_SE.index + 1 # Hacer que el índice comience en 1. 

# GRÁFICA SUBESTACIONES ----------------------------------------------------------------------------------------------------------------------------------------------------------
            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Recloser por SUBESTACION</p>", unsafe_allow_html=True)

            # Determinar el arreglo de las sub-figuras:
            n_fig_AMT=math.ceil(len(conteos_marcas_SE)/9) # Cantidad de subfiguras que tendran 9 SE c/u: Ejm: math.ceil(10/9)=2
            n_row_AMT=1 # Número de filas "SE"

            if n_fig_AMT>4: # Se tendrá un máximo de 4 columnas
                n_row_AMT=2 # Se tendrá un máximo de 2 filas

            n_column_AMT=n_fig_AMT//n_row_AMT # Cociente = n° de columnas

            # Dimensiones de cada gráfica:
            Ancho_AMT,height_AMT=8000,8000 # Dimensiones del ancho y la altura de los subplots de la figura: recloser por alimentador.
            Tamaño_FIG_SE=400

            l_column_widths_AMT=[Ancho_AMT]*n_column_AMT
            l_row_heights_AMT=[height_AMT]*n_row_AMT
            # Agregar los gráficos a cada subsubplot
            total_filas_grouped_aux = len(conteos_marcas_SE)
            cociente, residuo = divmod(total_filas_grouped_aux, n_fig_AMT)    
            #Obtener el nuevo residuo cuando el cociente se incrementa en 1
            cociente+=1
            residuo=total_filas_grouped_aux-cociente*n_fig_AMT
            #Creación de la lista donde se mostrarán las figuras
            lista_valores = [cociente] * n_fig_AMT
            lista_valores[-1] += residuo


            # Creación de la gráfica:
            indice_inicial = 0 # Variable que almacenará el última fila del grupo.
            figTOTAL  = sp.make_subplots( # Objeto que almacenará a las gráficas creadas por los plot
                                        rows=n_row_AMT,cols=n_column_AMT,
                                        column_widths=l_column_widths_AMT, row_heights=l_row_heights_AMT,
                                        vertical_spacing=0.2, horizontal_spacing=0.2)#, subplot_titles=['Subplot 1', 'Subplot 2', ...])
            for i in range(n_fig_AMT):
                
                # Creación del dataframe auxiliar
                limite_superior = indice_inicial + lista_valores[i]# Obtiene el límite superior del bloque actual
                df_FIG_SE = conteos_marcas_SE.iloc[indice_inicial:limite_superior].copy()# Crea un nuevo DataFrame copiando las filas correspondientes
                indice_inicial = limite_superior# Actualiza el índice inicial para el próximo ciclo

                # Graficar: El "go" tendrá un número de elementos igual al número de datos.
                traces=[
                    go.Bar(
                        name='Recloser instalados',
                        x=df_FIG_SE['Recloser instalados'],
                        y=df_FIG_SE['SUBESTACION'],
                        orientation='h',
                        marker_color="#C2BC18",
                        legendgroup='Recloser instalados'
                    ),
                    go.Bar(
                        name='Recloser con respuesta',
                        x=df_FIG_SE['Recloser con respuesta'],
                        y=df_FIG_SE['SUBESTACION'],
                        orientation='h',
                        marker_color="#e1b7ed",
                        legendgroup='Recloser con respuesta'
                    ),
                    go.Bar(
                        name='Recloser con comunicación',
                        x=df_FIG_SE['Recloser con comunicación'],
                        y=df_FIG_SE['SUBESTACION'],
                        orientation='h',
                        marker_color='#13B3C1',
                        legendgroup='Recloser con comunicación'
                    )
                ]
                fig1 = go.Figure(data=traces,layout=go.Layout(barmode='group')) # Agregar el grupo de cada trazo en la "fig1".
                fig1.update_traces(
                                    offset=-0.43, # Aumentar el alto de las barras.
                                    selector=dict(type='bar'),# Superponer las barras
                                    showlegend=False) # Ocultar la leyenda

                fig1.update_layout(
                    title="",
                    plot_bgcolor="rgba(0,0,0,0)",
                    yaxis_title='SUBESTACION',
                    xaxis_title='Cantidad de recloser',
                    title_x=0,
                    autosize=True,
                    height=200, #tamaño del gráfico
                    barmode='overlay' #superponer la barras en lugar de apilarlas
                )
                
                # fig1.update_xaxes(range=[0, df_FIG_SE['Recloser instalados'].max()+10],
                #                         tickvals=list(range(0, df_FIG_SE['Recloser instalados'].max()+20, 10)))

                fig1.update_traces(textposition='outside', textfont_size=5)
                

                # Posición (mxn) de la figura en el arreglo.
                row_idx = i // n_column_AMT + 1
                col_idx = i % n_column_AMT + 1 
                

                figTOTAL.add_trace(fig1['data'][0], row=row_idx, col=col_idx)
                figTOTAL.add_trace(fig1['data'][1], row=row_idx, col=col_idx)
                figTOTAL.add_trace(fig1['data'][2], row=row_idx, col=col_idx)
            
            figTOTAL.update_traces(showlegend=True)# Mostrar la leyenda (Los grupos creados)

            # MOSTRAR FIGURAS:
            # st.plotly_chart(figTOTAL, use_container_width=True) # Graficar en streamlit
            
            # with st.expander("UN_ViewData"):
            #     st.write(conteos_marcas_SE)
            #     download_excel(conteos_marcas_SE,"📥 Download Data","SUBESTACION-DATA.xlsx") # Descargar en formato (xlsx)

################################################################################################################################################################################
############################################### DISPOSICIÓN DE GRÁFICAS GRÁFICAS

        # i) DATOS GENERALES
            col1, col2, col3= st.columns([4,4,4]) # Poner los KPI's
            with col1:
                st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>BASE DE DATOS</p>", unsafe_allow_html=True)
                # with st.expander("ViewData ====> (Expandir)"):
                st.write(df_base_data) # Imprimir tabla         
                download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)
                st.markdown("---")

                st.markdown(f"<p style='font-size: 25px; text-align: Center; font-weight: bold;'>Recloser instalados: {total_recloser}</p>", unsafe_allow_html=True)
                # RECLOSER CON RESPUESTA Y COMUNICADOS
                st.markdown(f"<p style='font-size: 20px; text-align: Center'>\t Respuesta de recloser: {si_rpta_total}/{total_recloser}</p>", unsafe_allow_html=True)
                st.markdown(f"<p style='font-size: 20px;  text-align: Center'>\t Comunicación de recloser: {si_com_total}/{si_rpta_total}</p>", unsafe_allow_html=True)

                st.plotly_chart(figRC,use_container_width=True) # Gráfica del velocímetro
                st.markdown("---")
            with col2:
                    # DIAGRAMA DE BARRAS POR MARCAS
                    st.markdown(f"<p style='font-size: 18px; text-align: center; font-weight: bold;'>Recloser por marca</p>", unsafe_allow_html=True)
                    st.plotly_chart(figMARCA1, use_container_width=True, height=10) # Ancho del marco ocupa la columna de "col2"
                    st.markdown("---")
                    # DIAGRAMA DE PASTEL POR MARCAS
                    st.markdown(f"<p style='font-size: 18px; text-align: center; font-weight: bold;'>Porcentaje de recloser por marca</p>", unsafe_allow_html=True)
                    st.plotly_chart(figMARCA2,use_container_width=True, height=10) # Ancho del marco ocupa la columna de "col2"
                    # Tabla
                    st.write(conteos_marcas)
                    # st.write(conteos_marcas.style.background_gradient(cmap="Greens"))# Imprimir la tabla con escala de colores
                    download_excel(conteos_marcas,"📥 Download Data","Marca-DATA.xlsx") # Descargar en formato (xlsx)
            with col3:
                ######## GRÁFICA MAPA
                st.markdown(f"<p style='font-size: 18px; text-align: center; font-weight: bold;'>Recloser por departamento</p>", unsafe_allow_html=True)
                st.markdown(f"\t\t\t[[ Ubicación de recloser en Google Maps ]]({url_input})")# PONER URL DE LA UBICACIÓN DE LOS RECLOSER
                st_data = st_folium(map_folium, width=725)# Mostrar el mapa de Folium en Streamlit
                
                # Mostrar tabla
                with st.expander("DPTO_ViewData"):
                    st.write(conteos_marcas_DPTO)
                    download_excel(conteos_marcas_DPTO,"📥 Download Data","Departamento-DATA.xlsx") # Descargar en formato (xlsx)
                st.markdown("---") #separador

                ######## FIGURA DE UN (RADIO-CHART)
                st.markdown(f"<p style='font-size: 18px; text-align: Center'>Recloser por Unidad de Negocio</p>", unsafe_allow_html=True)            
                st.plotly_chart(fig_UN,use_container_width=True) # Ancho del marco ocupa la columna de "col3"

                with st.expander("UN_ViewData"):
                    st.write(conteos_marcas_UN)
                    download_excel(conteos_marcas_UN,"📥 Download Data","Unidad de Negocio-DATA.xlsx") # Descargar en formato (xlsx)

                st.markdown("---") #separador

    ######## FIGURA DE SUBESTACIONES
            st.markdown(f"<p style='font-size: 30x; text-align: Center'>Recloser por SUBESTACION</p>", unsafe_allow_html=True)

            st.plotly_chart(figTOTAL, use_container_width=True) # Graficar en streamlit

            with st.expander("UN_ViewData"):
                st.write(conteos_marcas_SE)
                download_excel(conteos_marcas_SE,"📥 Download Data","SUBESTACION-DATA.xlsx") # Descargar en formato (xlsx)
            st.markdown("---") #separador

            # **************************************
            # Estilo del "Streamlit"
            hide_st_style = """
                    <style>
        
                    footer {visibility: hidden;}

                    </style>
                    """

            st.markdown(hide_st_style, unsafe_allow_html= True)

        except Exception as e:
            st.write(e)
            st.error("...(Seleccionar los filtros)")

    elif selected_tab == lista_pestañas[2]:
        try:
            # 4° Periodo a elegir de "Registros.xlsx"
            workbook = openpyxl.load_workbook(name_excel)

            #4.1° Sacar todas las fechas a considerar
            st.sidebar.header("Periodo de consulta:")

            sheet_names = []# Lista que almacenará los nombres de las hojas
            for sheet in workbook.sheetnames:# Recorre todas las hojas del libro
                if not(sheet in ["SELECTORES","PLANTILLA","BDGeneral"]):
                    sheet_names.append(sheet) # Lista que almacenará los nombres de las hojas

            #4.2° Fecha de inicio
            hoja_excel = st.sidebar.selectbox("Fecha inicio:", sheet_names) # Usamos el widget selectbox para seleccionar una hoja
            for ii,vv in enumerate(sheet_names):
                if vv==hoja_excel:
                    item_inicio=ii

            #4.3° Fecha final
            hoja_excel_final = st.sidebar.selectbox("Fecha final:", sheet_names[item_inicio:][::-1])# Usamos el widget selectbox para seleccionar una hoja
            for ii,vv in enumerate(sheet_names):
                if vv==hoja_excel_final:
                    item_final=ii
            workbook.close()# Cierra el libro
            st.sidebar.markdown("----")

            # 5° Lectura de los datos del intervalor
            df = pd.read_excel(name_excel,sheet_name = hoja_excel_final)
                
            # 5° Creación de tablas
            #5.1 Creación de filtros
                # 5.1.1° Filtro de los campos
            dpto = df['DEPARTAMENTO'].unique()
            unidad_negocio = df['UNIDAD DE NEGOCIO'].unique()
            se = df['SUBESTACION'].unique()
            operador = df['OPERADOR INSTALADO'].unique()

                # 5.1.2° Filtro de Alimentador (AMT)
            st.sidebar.header("Filtro del Alimentador (AMT):")
                    # Forma 01: Ingresar desde el teclado
            # amt=st.sidebar.text_input("Escriba el alimentador (AMT):", "")

                    # Forma 02: Ingresar desde el teclado o seleccionar.
            amt_options = df[
                (df['DEPARTAMENTO'].isin(dpto)) &
                (df['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                (df['SUBESTACION'].isin(se)) &
                (df['OPERADOR INSTALADO'].isin(operador))
            ]['AMT'].unique()
            
            lista_amt_options=sorted(amt_options.tolist())
            amt_input = st.sidebar.text_input("Escriba el alimentador (AMT):", "")
            filtered_amt_options = [option for option in lista_amt_options if amt_input.lower() in option.lower()]
            amt = st.sidebar.selectbox("Seleccione el alimentador (AMT):", options=filtered_amt_options, index=0 if filtered_amt_options else None)

                # 5.1.3° Filtro del recloser por código: "SCADA" o "GIS"
            st.sidebar.markdown("----")
            filtro_opcion = st.sidebar.selectbox("Elección del campo:", ["1- SCADA", "2- GIS"])# Barra desplegable para seleccionar entre "SCADA" y "GIS"
            if filtro_opcion == "1- SCADA":
                field_camp = 'Codigo SCADA Actual'

                st.sidebar.header("Seleccione el código SCADA:")
                scada_options = df[df['AMT'] == amt]['Codigo SCADA Actual'].unique() # Filtrar las opciones
                scada_options = sorted(scada_options) # Ordenar las opciones

                scada_options_non_empty = [option for option in scada_options if option.strip() != '']# Filtra los elementos que contienen "RE"
                seleccion = st.sidebar.radio("", options= scada_options_non_empty, index=0) # Muestrame en el cuadro de los check-list
            
            elif filtro_opcion == "2- GIS":
                field_camp = 'SECC. GIS NUEVO'
                st.sidebar.header("Seleccione el código del SECCIONAMIENTO")
                gis_options = df[df['AMT'] == amt]['SECC. GIS NUEVO'].unique()
                gis_options = sorted(gis_options)

                gis_options_filtered = [option for option in gis_options if option.strip() != '' and option != '--']
                seleccion = st.sidebar.radio("", options=list(gis_options), index=0)
                
            # 5.2° Creación de la base de datos
            diccionario_fechas={
                "Fecha":[],
                "Nro de respuestas":[],
                "Nro de intermitencias":[],
                "Nro de muestras":[],
                "Rpta actual":[],
                "Comunicación actual":[]
            }
            
            for hoja_recorrido in sheet_names[item_inicio:item_final+1]:
                # 5.2.1° Obtención de los datos por fechas del recloser seleccionado.
                df_aux = pd.read_excel(name_excel,sheet_name = hoja_recorrido)
                
                diccionario_fechas['Fecha'].append(hoja_recorrido)
                diccionario_fechas['Nro de respuestas'].append(df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Nro de respuestas'].iloc[0])
                        # Definición del diccionario:
                            # df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Nro de respuestas'].iloc[0]:
                            # Mascara: (df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion)
                            # Valor que deseamos sacar: df_aux.loc[... , "field_name" ] => Como un dataframe
                            # Solo obtener el valor: .iloc[0] => Es solo un número.
                
                diccionario_fechas['Nro de intermitencias'].append(df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Nro de intermitencias'].iloc[0])
                diccionario_fechas['Nro de muestras'].append(df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Nro de muestras'].iloc[0])
                diccionario_fechas['Rpta actual'].append(df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Rpta actual'].iloc[0])
                diccionario_fechas['Comunicación actual'].append(df_aux.loc[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion), 'Comunicación actual'].iloc[0])


                # 5.2.2° Obtención de todos los datos, del recloser seleccionado
            df_fila_seleccionada = df[(df_aux['AMT'] == amt) & (df_aux[field_camp] == seleccion)].iloc[0,:13].to_frame().T #Datos de la fila que se selecciona.
                #iloc[0,:13]: Mantenerme las 13 primeras columnas.
                #.to_frame().T: Transponición de filas y columnas.
            df_fila_seleccionada = pd.concat([df_fila_seleccionada] * (item_final-item_inicio+1), ignore_index=True)

                # 5.2.3° Aumentar las filas de fila_seleccionada y poner los valores de la primera fila en las filas nuevas.
            df_final=pd.DataFrame(diccionario_fechas)
            df_final = pd.concat(
                [df_final.iloc[:, :1], df_fila_seleccionada.iloc[:, :], df_final.iloc[:, 1:].reset_index(drop=True)],
                 axis=1) # Ordenar: 1er columna(df_final)+todas las columnas(df_fila_seleccionada)+2da hacia delante las columnas(df_final)

            # 5.3° Filtro de los campos
            # Mostrar la base de datos
            st.sidebar.markdown("---")
            st.sidebar.header("BASE DE DATOS")
            selected_columns_difference=[col for col in df_final.columns if col not in selected_columns_comunicacion] # Muestra de los demás campos.
            
            select_all_field = st.sidebar.checkbox("Mostrar todos los CAMPOS.") # Casilla de verificación para seleccionar todos los filtros.
            lista_field=sorted(selected_columns_difference) # Lista que tiene a las opciones ordenadas.

            if select_all_field:
                st.sidebar.warning("Se mostrará todos los CAMPOS de la Base de Datos.")
                fields = lista_field

            else:
                fields = st.sidebar.multiselect(
                    "Filtrar campos",
                    options=lista_field,
                    default=[],
                )
            st.sidebar.markdown("---") #separador
            
            selected_columns=selected_columns_comunicacion+fields # Lista con los campos a mostrar.
            df_base_data = df_final[df_final.columns.intersection(selected_columns)]# Ordenar las columnas del Dataframe en base al orden de una lista.
            df_base_data = df_base_data.set_index(df_base_data.columns[0]) #Convertir la primera columna en el índice del dataframe.

            # st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>BASE DE DATOS</p>", unsafe_allow_html=True)
            # with st.expander("ViewData ====> (Expandir)"):
            # st.write(df_base_data) # Imprimir tabla         
            # download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)
            # st.markdown("---")

#**********************************************************************************************************************************************************
# 6° GRÁFICA RESPUESTA/COMUNICACIÓN ----------------------------------------------------------------------------------------------------------------------------------------------------------
            # 6.1° Adaptación del dataframe.
            df_grafico = df_final[["Fecha", "Rpta actual", "Comunicación actual"]].copy() # Copiar solo 3 campos del df_final.
            df_grafico.replace({'Si': 2, 'No': 1}, inplace=True) # Reemplazar los valores "Si" y "No" por 2 y 1

            df_grafico['Color_rpta'] = df_grafico['Rpta actual'].map({2: si_rpta_color, 1: no_color}) # Crear nueva columna para "Asignar colores"
            df_grafico['Texto_rpta'] = df_grafico['Rpta actual'].map({2: 'Si', 1: 'No'}) # Crear nueva columna para "Asignar texto", encima de las barras

            df_grafico['Color_com'] = df_grafico['Comunicación actual'].map({2: si_com_color, 1: no_color}) # Crear nueva columna para "Asignar colores"
            df_grafico['Texto_com'] = df_grafico['Comunicación actual'].map({2: 'Si', 1: 'No'}) # Crear nueva columna para "Asignar texto", encima de las barras


            # 6.2° Crear diccionario para el color de las barras
            diccionario_color_res={}
            diccionario_color_com={}

            for index,row in df_grafico.iterrows():
                ff,color_rpta,color_com=row['Fecha'],row['Color_rpta'],row['Color_com']
                diccionario_color_res[ff]=color_rpta
                diccionario_color_com[ff]=color_com

            # 6.2° Evolución de la respuesta en el tiempo
                # 6.2.1° Crear el gráfico
            figRES =  px.bar(df_grafico, x="Fecha", y="Rpta actual",
                        color="Fecha",
                        color_discrete_map=diccionario_color_res, # Colores en base al diccionario
                        text=df_grafico['Texto_rpta'])

                # 6.2.2° Configurar el gráfico
            figRES.update_layout(xaxis=dict(tickangle=-45, tickfont=dict(size=15)),
                            yaxis=dict(showticklabels=False, range=[0, 2.5]))

            figRES.update_traces(textposition='outside',
                            textfont_size=15,
                            showlegend=False) # Ocultar la legenda.

            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Respuesta de recloser</p>", unsafe_allow_html=True)  
            # st.plotly_chart(figRES, use_container_width=True, height=200)# Mostrar el gráfico en Streamlit

            # 6.3° Evolución de la comunicación en el tiempo
                # 6.3.1° Crear el gráfico
            figCOM =  px.bar(df_grafico, x="Fecha", y="Comunicación actual",
                        color="Fecha",
                        color_discrete_map=diccionario_color_com, # Colores en base al diccionario
                        text=df_grafico['Texto_com'])

                # 6.3.2° Configurar el gráfico
            figCOM.update_layout(xaxis=dict(tickangle=-45, tickfont=dict(size=15)),
                            yaxis=dict(showticklabels=False, range=[0, 2.5]))

            figCOM.update_traces(textposition='outside',
                            textfont_size=15,
                            showlegend=False) # Ocultar la legenda.
            
            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Comunicación de recloser</p>", unsafe_allow_html=True)
            # st.plotly_chart(figCOM, use_container_width=True, height=200)# Mostrar el gráfico en Streamlit

#**********************************************************************************************************************************************************
# 7° GRÁFICA NRO RESPUESTAS/INTERMITENCIAS ----------------------------------------------------------------------------------------------------------------------------------------------------------
            # 7.1° Adaptación al dataframe.

            df_grafico = df_final[["Fecha", "Nro de respuestas", "Nro de intermitencias"]].copy() # Copiar solo 3 campos del df_final.

                # Crear el gráfico de líneas con Plotly Express
            figNRO = px.line()

                # Agregar líneas punteadas hacia arriba en cada fecha
            for date in df_grafico['Fecha']:
                figNRO.add_shape(
                    type='line',
                    x0=date,
                    x1=date,
                    y0=-2,
                    y1=df_grafico['Nro de respuestas'].max()+2,
                    line=dict(dash='dash', color='#FFFF93'),
                    name='Nro de respuestas'
                )
            
                # Añadir otra línea para 'Nro de respuestas'
            figNRO.add_trace(
                go.Scatter(
                    x=df_grafico['Fecha'],
                    y=df_grafico['Nro de respuestas'],
                    mode='lines+markers',  # Incluir línea y puntos
                    line=dict(color='blue'), # Color de la línea
                    marker=dict(size=10), # Tamaño de los puntos
                    name='Nro de respuestas'
                )
            )
                # Añadir otra línea para 'Nro de intermitencias'
            figNRO.add_trace(
                go.Scatter(
                    x=df_grafico['Fecha'],
                    y=df_grafico['Nro de intermitencias'],
                    mode='lines+markers',  # Incluir línea y puntos
                    line=dict(color='red'), # Color de la línea
                    marker=dict(size=10), # Tamaño de los puntos
                    name='Nro de intermitencias'
                )
            )
                # Configurar las gráficas
            figNRO.update_layout(xaxis=dict(tickangle=-45, tickfont=dict(size=15)),
                            yaxis=dict(tickfont=dict(size=15),dtick=3))

                
            # st.markdown(f"<p style='font-size: 22px; text-align: Center'>Intermitencias producidas (MUESTRA=15)</p>", unsafe_allow_html=True)
            # st.plotly_chart(figNRO,use_container_width=True, height=200) # Mostrar el gráfico en Streamlit

################################################################################################################################################################################
############################################### DISPOSICIÓN DE GRÁFICAS GRÁFICAS

        # i) DATOS GENERALES
            col1, col2= st.columns([4,3])
            with col1:
                st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>BASE DE DATOS</p>", unsafe_allow_html=True)
                # with st.expander("ViewData ====> (Expandir)"):
                st.write(df_base_data) # Imprimir tabla         
                download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)
                st.markdown("---")

                st.markdown(f"<p style='font-size: 22px; text-align: Center'>Intermitencias producidas (MUESTRA=15)</p>", unsafe_allow_html=True)
                st.plotly_chart(figNRO,use_container_width=True, height=100) # Mostrar el gráfico en Streamlit
            with col2:
                st.markdown(f"<p style='font-size: 22px; text-align: Center'>Respuesta de recloser</p>", unsafe_allow_html=True)  
                st.plotly_chart(figRES, use_container_width=True, height=100)# Mostrar el gráfico en Streamlit

                st.markdown(f"<p style='font-size: 22px; text-align: Center'>Comunicación de recloser</p>", unsafe_allow_html=True)
                st.plotly_chart(figCOM, use_container_width=True, height=100)# Mostrar el gráfico en Streamlit
            
        except Exception as e:
            st.error("...(Escriba correctamente el alimentador-AMT)")

    elif selected_tab==lista_pestañas[3]:
        try:
            # st.header("¡UPS!, Esta pestaña se encuentra en actualizacion.")
            # 4° Periodo a elegir de "Registros.xlsx"
            workbook = openpyxl.load_workbook(name_excel)

            #4.1° Sacar todas las fechas a considerar
            st.sidebar.header("Periodo de consulta:")

            sheet_names = []# Lista que almacenará los nombres de las hojas
            for sheet in workbook.sheetnames:# Recorre todas las hojas del libro
                if not(sheet in ["SELECTORES","PLANTILLA","BDGeneral"]):
                    sheet_names.append(sheet) # Lista que almacenará los nombres de las hojas

            #4.2° Fecha de inicio
            hoja_excel = st.sidebar.selectbox("Fecha inicio:", sheet_names) # Usamos el widget selectbox para seleccionar una hoja
            for ii,vv in enumerate(sheet_names):
                if vv==hoja_excel:
                    item_inicio=ii

            #4.3° Fecha final
            hoja_excel_final = st.sidebar.selectbox("Fecha final:", sheet_names[item_inicio:][::-1])# Usamos el widget selectbox para seleccionar una hoja
            for ii,vv in enumerate(sheet_names):
                if vv==hoja_excel_final:
                    item_final=ii
            workbook.close()# Cierra el libro
            st.sidebar.markdown("----")

            # 5° Lectura de los datos del intervalor
            df = pd.read_excel(name_excel,sheet_name = hoja_excel_final)

            #5.1 Creación de filtros
                # 5.1.1° Filtro de los campos
            dpto = df['DEPARTAMENTO'].unique()

            # Filtro de Unidad de Negocio (se habilita en función de la selección de Departamento)
            st.sidebar.header("Unidad de Negocio:")
            select_all_UN= st.sidebar.checkbox("Marcar todas las UN.")#Casilla de verificación para seleccionar todos los filtros.
            un_options = df[df['DEPARTAMENTO'].isin(dpto)]['UNIDAD DE NEGOCIO'].unique() # Filtro las UN por los dptos seleccionados.
            lista_UN_options=sorted(un_options.tolist())

            if select_all_UN:
                st.sidebar.warning("Seleccionaste todas las UN.")
                unidad_negocio = lista_UN_options
            else:
                unidad_negocio = st.sidebar.selectbox(
                    "Filtrar UN:",
                    options=lista_UN_options, # Opciones, incluido a las UN ordenadas.
                    default=[],
                )

            # Filtro de Subestación (se habilita en función de la selección de Unidad de Negocio)
            st.sidebar.header("Subestación Eléctrica")
            select_all_SE= st.sidebar.checkbox("Marcar todas las SE.")#Casilla de verificación para seleccionar todos los filtros.
            se_options = df[
                (df['DEPARTAMENTO'].isin(dpto)) &
                (df['UNIDAD DE NEGOCIO'].isin(unidad_negocio))
                ]['SUBESTACION'].unique()
            lista_se_options=sorted(se_options.tolist())

            if select_all_SE:
                st.sidebar.warning("Seleccionaste todas las SE.")
                se = lista_se_options
            else:
                se = st.sidebar.multiselect(
                    "Filtrar SE:",
                    options=lista_se_options,
                    default=[],
                )
                
            # Filtro de Operador (se habilita en función de la selección de Subestación)
            operador_options = df[
                (df['DEPARTAMENTO'].isin(dpto)) &
                (df['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                (df['SUBESTACION'].isin(se))
            ]['OPERADOR INSTALADO'].unique()
            operador=sorted(operador_options.tolist())
                
            # Filtro de Alimentador (AMT) - Filtrado en base a todas las selecciones anteriores
            amt_options = df[
                (df['DEPARTAMENTO'].isin(dpto)) &
                (df['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                (df['SUBESTACION'].isin(se)) &
                (df['OPERADOR INSTALADO'].isin(operador))
            ]['AMT'].unique()
            amt=sorted(amt_options.tolist())
            
            # 5.2° Filtra el DataFrame en función de las selecciones:
            filtered_df = df[
                (df['DEPARTAMENTO'].isin(dpto)) &
                (df['UNIDAD DE NEGOCIO'].isin(unidad_negocio)) &
                (df['SUBESTACION'].isin(se)) &
                (df['OPERADOR INSTALADO'].isin(operador)) &
                (df['AMT'].isin(amt))
            ]

            # 5.3° Filtrar los recloser:
                # Filtrar Dataframe
            df_filtro_Marca=filtered_df[filtered_df['MARCA'].isin(lista_recloser)]
            df_filtro_Marca["MARCA"]=df_filtro_Marca["MARCA"].replace("NOJA Power", "NOJA")   
                                                                                                                #El "\" indica un salto.
            condicion_filtrar=((df_filtro_Marca["MARCA"] == "S&C") & (df_filtro_Marca["SECC. GIS NUEVO"] == "--")) | \
                ((df_filtro_Marca['MARCA'] == 'ABB') & (df_filtro_Marca['Controlador'] != "PCD2000R")) | \
                ((df_filtro_Marca['MARCA'] == 'SEL') & (df_filtro_Marca['Controlador'] != "SEL-351R"))
            
            df_filtro_Marca = df_filtro_Marca[~condicion_filtrar] # Eliminar los registros que cumplan con la condición.
            df_filtro_Marca = df_filtro_Marca.reset_index(drop=True) # Dataframe filtrado, que se usará para los siguientes filtros.

                # Eliminar todas las filas que tienen "--" en la columna "Comunicación actual"
            df_filtro_Marca = df_filtro_Marca[df_filtro_Marca['Comunicación actual'] != '--']
            df_filtro_Marca.reset_index(drop=True, inplace=True)# Restablecer el índice después de eliminar las filas
            df_filtro_Marca['Codigo SCADA Actual'] = df_filtro_Marca['Codigo SCADA Actual'].fillna('vacío')# Cambiar todos los elementos None en la columna "Codigo SCADA Actual".

            List_ptos_AMT,List_ptos_REC=df_filtro_Marca["AMT"].tolist(),df_filtro_Marca["Codigo SCADA Actual"]

            #6° Ingreso de número de puntos críticos
            maximo=len(List_ptos_AMT)
            if maximo>20:
                maximo=20
            num_total_ptos = st.sidebar.slider("Número de puntos críticos a mostrar: ", min_value=1, max_value=maximo, step=1)

            # 7° Conteo de los "NO" de cada recloser.
            List_conteo_NO_ptos=[]
            for contador in range(0,len(List_ptos_AMT)):
                AMT_pto,SCADA_pto=List_ptos_AMT[contador],List_ptos_REC[contador]
                kcount_com=0 #Contador que determina los días que no hubo comunicación.

                for hoja_recorrido in sheet_names[item_inicio:item_final+1]:
                    df_aux = pd.read_excel(name_excel,sheet_name = hoja_recorrido)

                    mascara_pto=(df_aux["AMT"] == AMT_pto) & (df_aux["Codigo SCADA Actual"] == SCADA_pto) # Creación de la máscara.
                    estado_comunicacion=df_aux.loc[mascara_pto, "Comunicación actual"].iloc[0] # Localización del valor del otro campo.

                    if estado_comunicacion=="No":
                        kcount_com+=1
                    
                List_conteo_NO_ptos.append(kcount_com)

            # 8° Crear el DataFrame de los conteos "NO"
            df_conteo_NO = pd.DataFrame({
                'AMT': List_ptos_AMT,
                'Codigo SCADA Actual': List_ptos_REC,
                'N° días incomunicados': List_conteo_NO_ptos
            })
            
            df_conteo_NO=df_conteo_NO.sort_values(by='N° días incomunicados',ascending=False).reset_index(drop=True) # Ordenar y reiniciar el índice.
            # 9° Graficar
            df_graficas_ptos = df_conteo_NO.head(num_total_ptos) # Filtrar los "n" puntos críticos
            fig_total_pto=[] # Objeto que almacenará todas las figuras a graficar.
            lista_nro_dias_incomunicado_BD=[] # Objeto que almacenará el total de días incomunicados por los recloser.

            for index,row in df_graficas_ptos.iterrows():
                AMT_pto,SCADA_pto=row["AMT"],row["Codigo SCADA Actual"]
                diccionario_fechas={
                    "Fecha":[],
                    "Comunicación actual":[]
                }

                for hoja_recorrido in sheet_names[item_inicio:item_final+1]:
                # 9.1° Obtención de los datos por fechas del recloser seleccionado.
                    df_aux = pd.read_excel(name_excel,sheet_name = hoja_recorrido)
                    diccionario_fechas['Fecha'].append(hoja_recorrido)
                    diccionario_fechas['Comunicación actual'].append(df_aux.loc[(df_aux['AMT'] == AMT_pto) & (df_aux["Codigo SCADA Actual"] == SCADA_pto), 'Comunicación actual'].iloc[0])
                    if hoja_recorrido==sheet_names[item_inicio]:
                        df_data=df_aux.loc[(df_aux['AMT'] == AMT_pto) & (df_aux["Codigo SCADA Actual"] == SCADA_pto)] # Copiar la data del recloser "k" considerado como punto crítico.
                        
                        lista_nro_dias_incomunicado_BD.append(df_graficas_ptos.loc[index, "N° días incomunicados"]) # Almacenar en una lista los valores de nro de días incomunicados
                        df_data=df_data.drop(columns=['Nro de respuestas','Nro de intermitencias','Nro de muestras','Rpta actual','Comunicación actual'])
                    
                # 9.2° Creación de la base de datos
                if index==0: #Crear el dataframe que contendrá todos los elementos
                    df_final = pd.DataFrame(columns=df_data.columns)
                # Concatenar los puntos críticos
                dfs_a_concatenar = []# Lista para almacenar los DataFrames a concatenar
                for index, row in df_data.iterrows():# Creación de una lista de DF.
                    dfs_a_concatenar.append(pd.DataFrame(row).T)
                
                df_aux_pto = pd.concat(dfs_a_concatenar, ignore_index=True)# Concatener todos los DF de la lista en uno solo.
                df_final = pd.concat([df_final, df_aux_pto], ignore_index=True) # Concatenar los dos dataframes.
                df_final["N° días incomunicados"]=lista_nro_dias_incomunicado_BD # Poner la columna que nos da el nro de días que el recloser estuvo incomunicado.

                # 9.3° Adaptación del dataframe.
                df_grafico = pd.DataFrame(diccionario_fechas) # Crear un dataframe
                df_grafico.replace({'Si': 2, 'No': 1}, inplace=True) # Reemplazar los valores "Si" y "No" por 2 y 1

                df_grafico['Color_com'] = df_grafico['Comunicación actual'].map({2: si_com_color, 1: no_color}) # Crear nueva columna para "Asignar colores"
                df_grafico['Texto_com'] = df_grafico['Comunicación actual'].map({2: 'Si', 1: 'No'}) # Crear nueva columna para "Asignar texto", encima de las barras

                # 9.4° Crear diccionario para el color de las barras
                diccionario_color_com={}

                for index,row in df_grafico.iterrows():
                    ff,color_com=row['Fecha'],row['Color_com']
                    diccionario_color_com[ff]=color_com

                # 9.5° Evolución de la comunicación en el tiempo
                    # 9.5.1° Crear el gráfico
                figCOM =  px.bar(df_grafico, x="Fecha", y="Comunicación actual",
                            color="Fecha",
                            color_discrete_map=diccionario_color_com, # Colores en base al diccionario
                            text=df_grafico['Texto_com'])

                    # 9.5.2° Configurar el gráfico
                figCOM.update_layout(xaxis=dict(tickangle=-45, tickfont=dict(size=15)),
                                yaxis=dict(showticklabels=False, range=[0, 2.5]))

                figCOM.update_traces(textposition='outside',
                                textfont_size=15,
                                showlegend=False) # Ocultar la legenda.
                
                fig_total_pto.append(figCOM)
            


            # 10° Filtro de los campos
            # Mostrar la base de datos
            st.sidebar.markdown("---")
            st.sidebar.header("BASE DE DATOS")
            selected_columns_difference=[col for col in df_final.columns if col not in selected_columns_criticos] # Muestra de los demás campos.
            
            select_all_field = st.sidebar.checkbox("Mostrar todos los CAMPOS.") # Casilla de verificación para seleccionar todos los filtros.
            lista_field=sorted(selected_columns_difference) # Lista que tiene a las opciones ordenadas.

            if select_all_field:
                st.sidebar.warning("Se mostrará todos los CAMPOS de la Base de Datos.")
                fields = lista_field

            else:
                fields = st.sidebar.multiselect(
                    "Filtrar campos",
                    options=lista_field,
                    default=[],
                )
            st.sidebar.markdown("---") #separador
            
            selected_columns=selected_columns_criticos+fields # Lista con los campos a mostrar.
            df_base_data = df_final[df_final.columns.intersection(selected_columns)]# Ordenar las columnas del Dataframe en base al orden de una lista.
            df_base_data.index = df_base_data.index + 1 # Hacer que la llave primaria inicie en "1" (INDEX).

            # # Mostrar la tabla
            # st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>BASE DE DATOS</p>", unsafe_allow_html=True)
            # st.write(df_base_data)
            # download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)
            # st.markdown("---")
            
################################################################################################################################################################################
############################################### DISPOSICIÓN DE GRÁFICAS GRÁFICAS
            
            # 10° Mostrar la tabla de la base de datos:
            st.markdown(f"<p style='font-size: 20px; text-align: center; font-weight: bold;'>BASE DE DATOS</p>", unsafe_allow_html=True)
            st.write(df_base_data)
            download_excel(df_base_data,"📥 Download Data","Tabla_Datos-DATA.xlsx") # Descargar en formato (xlsx)
            
            st.markdown("---")
            st.markdown(f"<p style='font-size: 25px; text-align: center; font-weight: bold;'>Comunicación de {num_total_ptos}/{len(List_ptos_AMT)} recloser</p>", unsafe_allow_html=True)
            # 11° Graficar en streamlit
            for fig_count in range(0,num_total_ptos,3):
                try:
                    col1,col2,col3=st.columns([4,4,4])
                    with col1:
                        AMT_pto,SCADA_pto=df_base_data.loc[fig_count+1,"AMT"],df_base_data.loc[fig_count+1,"Codigo SCADA Actual"]
                        st.plotly_chart(fig_total_pto[fig_count], use_container_width=True, height=50)# Mostrar el gráfico en Streamlit
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Punto crítico N°{fig_count+1}:</p>", unsafe_allow_html=True)
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Recloser {AMT_pto}_{SCADA_pto}</p>", unsafe_allow_html=True)

                    
                    with col2:
                        AMT_pto,SCADA_pto=df_base_data.loc[fig_count+2,"AMT"],df_base_data.loc[fig_count+2,"Codigo SCADA Actual"]
                        st.plotly_chart(fig_total_pto[fig_count+1], use_container_width=True, height=50)# Mostrar el gráfico en Streamlit
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Punto crítico N°{fig_count+2}:</p>", unsafe_allow_html=True)
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Recloser {AMT_pto}_{SCADA_pto}</p>", unsafe_allow_html=True)

                    with col3:
                        AMT_pto,SCADA_pto=df_base_data.loc[fig_count+3,"AMT"],df_base_data.loc[fig_count+3,"Codigo SCADA Actual"]
                        st.plotly_chart(fig_total_pto[fig_count+2], use_container_width=True, height=50)# Mostrar el gráfico en Streamlit
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Punto crítico N°{fig_count+3}:</p>", unsafe_allow_html=True)
                        st.markdown(f"<p style='font-size: 18px; text-align: Center'>Recloser {AMT_pto}_{SCADA_pto}</p>", unsafe_allow_html=True)

                except Exception as e:
                    sys.exit() # En caso de existir un error. Terminar de ejecutar el programa.

        except Exception as e:
            st.error("...(Seleccionar los filtros)")